import matplotlib.pyplot as plt
import numpy as np
import openpyxl
print('Автор: Василь Онуфрійчук') 
print('Побудова графіка по точкам із файлу ексель.')
print('')
x=list()
y=list()
table_name='Таблиця з даними для побудови графіка.xlsx'
def build_list(a,b,h):
    k=2
    for i in range(h):
        x=sheet_wb[b+str(k)].value
        a.append(x)
        k+=1
def choise_grid(a):
    if a==1:
        g='x'
    elif a==2:
        g='y'
    elif a==3:
        g='both'
    else:
        g=False
    return g
try:
    wb = openpyxl.load_workbook(table_name)
    sheet_wb=wb['Дані для побудови графіка']
    c_wb=sheet_wb['C2'].value
    d_wb=sheet_wb['D2'].value
    e_wb=sheet_wb['E2'].value
    f_wb=sheet_wb['F2'].value
    g_wb=sheet_wb['G2'].value
    h_wb=sheet_wb['H2'].value
    build_list(x,'A',c_wb)
    build_list(y,'B',c_wb)
    myhex = str(g_wb)
    fig, ax = plt.subplots()
    plt.title(str(d_wb))
    plt.xlabel(str(e_wb))
    plt.ylabel(str(f_wb))
    ax.plot(x, y, color=myhex)
    if h_wb==1 or h_wb==2 or h_wb==3:
        ax.grid(axis=choise_grid(h_wb))
    plt.show()
except:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Дані для побудови графіка'
    ws['A1'] = 'x'
    ws['B1'] = 'y'
    ws['C1'] = 'Кількість точок(ціле число)'
    ws['C2'] = 0
    ws['D1'] = 'Заголовок'
    ws['D2'] = 'Мій графік'
    ws['E1'] = 'Текст осі х'
    ws['E2'] = 'Вісь х'
    ws['F1'] = 'Текст осі у'
    ws['F2'] = 'Вісь у'
    ws['G1'] = 'Колір лінії(hex)'
    ws['G2'] = '#32a852'
    ws['H1'] = 'Сітка (1-вертикальна 2-горизонтальна 3-вертикальна+горизонтальна)'
    ws['H2'] = 0
    wb.save(table_name)


